# -*- coding: utf-8 -*-
from glob import glob
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Style
from openpyxl.styles.colors import YELLOW

gul_stil = Style(fill=PatternFill(patternType='solid', fgColor=YELLOW))

# finn alle filene som passer med filnavn
path = r"C:\Users\Yngvild Gagnat\Desktop\BEV3901\Automatisk SPSS\ut"
path += "\\"
filnavn = path + r"*_normality_log_Pt.xlsx"
filer = glob(filnavn)

if not filer:
    print('Fant ingen filer med navn %s' % filnavn)

# variabel hvor vi skal lagre det vi henter ut
normalities = []

# gjør dette for alle filene
for fil in filer:
    # last inn excel dokumentet
    bok = load_workbook(fil)
    # Velg Sheet1
    ark = bok.get_sheet_by_name("Sheet1")

    for rad in ark.rows:
        forste_celle = rad[0]
        if forste_celle.value == 'Tests of Normality':
            rad_nummer = forste_celle.row + 3
            break

    # hent navnet
    navn = ark.cell(row=rad_nummer, column=1).value
    # hent verdien
    normality = ark.cell(row=rad_nummer, column=7).value

    # sjekk om vi fant navnet og verdien
    if navn == None or normality == None:
        print("Fant ikke normality i", fil)
        continue

    # legg til navn og normality som et par i listen normalities
    normalities.append((navn, normality))
    print(".", end="", flush=True)

# lag en ny excel fil
ut_bok = Workbook()
# lag et ark
ark = ut_bok.get_active_sheet()

# for alle oppføringer i normalities
for n, norm in enumerate(normalities):
    navn_celle = ark.cell(row=n+1, column=1)
    verdi_celle = ark.cell(row=n+1, column=2)

    # hver oppføring i normalities er lagret som par nummerert 0 og 1
    navn_celle.value = norm[0]
    verdi_celle.value = norm[1]

    if norm[1] >= 0.05:
        #navn_celle.font = Font(bold=True)
        #verdi_celle.font = Font(bold=True)
        navn_celle.style = gul_stil
        verdi_celle.style = gul_stil


ut_bok.save(path + r'normalities_log_Pt.xlsx')
print("ferdig")
