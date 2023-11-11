from glob import glob
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Style
from openpyxl.styles.colors import YELLOW

gul_stil = Style(fill=PatternFill(patternType='solid', fgColor=YELLOW))

# finn alle filene som passer med filnavn
path = r"C:\Users\Yngvild Gagnat\Desktop\BEV3901\Automatisk SPSS\ut"
path += '\\' # slash på slutten av path
# path = 'spss/' # på arves maskin
filnavn = path + r"*_normality_pot.xlsx"
filer = glob(filnavn)

# variabel hvor vi skal lagre det vi henter ut
normalities = []

# gjør dette for alle filene
for fil in filer:
    if '~$' in fil:
        print('hopper over', fil)
        continue
    # last inn excel dokumentet
    bok = load_workbook(fil)
    # Velg Sheet1
    ark = bok.get_sheet_by_name("Sheet1")

    for rad in ark.rows:
        forste_celle = rad[0]
        if forste_celle.value == 'Tests of Normality':
            rad_nummer = forste_celle.row + 3
            break

    # hent ut maks 32 verdier, avbryt når vi finner "*. This is a lower..."
    navn_verdier = ['x', 'x', 'x', 'x']
    for r in range(32):
        navn_celler = (
            ark.cell(row=rad_nummer + r, column=4), # dag
            ark.cell(row=rad_nummer + r, column=2), # leg
            ark.cell(row=rad_nummer + r, column=3), # pre/post
            ark.cell(row=rad_nummer + r, column=1)) # test

        # bruk forrige verdi hvis verdien ikke eksisterer
        for i,celle in enumerate(navn_celler):
            if celle.value != None:
                navn_verdier[i] = celle.value

        navn = '_'.join(navn_verdier)
        normality = ark.cell(row=rad_nummer + r, column=10).value

        if normality == None:
            print('Hentet %s verdier fra %s' %(r, fil))
            break

        # sjekk om vi fant verdien
        if type(normality) != float:
            print("Feil med normality i %s på rad %s" % (fil, rad_nummer + r))
            continue

        # legg til navn og normality som et par i listen normalities
        normalities.append((navn, normality))

# lag en ny excel fil
ut_bok = Workbook()
# lag et ark
ark = ut_bok.get_active_sheet()

# for alle oppføringer i normalities
for n, norm in enumerate(normalities):
    navn_celle = ark.cell(row=n+1, column=1)
    verdi_celle = ark.cell(row=n+1, column=2)

    # hver oppføring i normalities er lagret som par nummerert 0 og 1
    navn_celle.value = norm[0]
    verdi_celle.value = norm[1]

    if norm[1] >= 0.05:
        #navn_celle.font = Font(bold=True)
        #verdi_celle.font = Font(bold=True)
        navn_celle.style = gul_stil
        verdi_celle.style = gul_stil


ut_bok.save(path + r'normalities_pot.xlsx')
print("ferdig")
